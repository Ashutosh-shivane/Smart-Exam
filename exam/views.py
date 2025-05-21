from django.shortcuts import render,redirect,reverse
from . import forms,models
from django.db.models import Sum
from calendar import month_name
from django.contrib.auth.models import Group
from django.http import HttpResponseRedirect
from django.contrib.auth.decorators import login_required,user_passes_test
from django.conf import settings
from datetime import date, timedelta
from django.db.models import Q
from django.core.mail import send_mail
from teacher import models as TMODEL
from student import models as SMODEL
from teacher import forms as TFORM
from student import forms as SFORM
from django.contrib.auth.models import User



def home_view(request):
    if request.user.is_authenticated:
        return HttpResponseRedirect('afterlogin')  
    return render(request,'exam/index.html')


def is_teacher(user):
    return user.groups.filter(name='TEACHER').exists()

def is_student(user):
    return user.groups.filter(name='STUDENT').exists()

def afterlogin_view(request):
    if is_student(request.user):      
        return redirect('student/student-dashboard')
                
    elif is_teacher(request.user):
        accountapproval=TMODEL.Teacher.objects.all().filter(user_id=request.user.id,status=True)
        if accountapproval:
            return redirect('teacher/teacher-dashboard')
        else:
            return render(request,'teacher/teacher_wait_for_approval.html')
    else:
        return redirect('admin-dashboard')



def adminclick_view(request):
    if request.user.is_authenticated:
        return HttpResponseRedirect('afterlogin')
    return HttpResponseRedirect('adminlogin')


@login_required(login_url='adminlogin')
def admin_dashboard_view(request):
    dict={
    'total_student':SMODEL.Student.objects.all().count(),
    'total_teacher':TMODEL.Teacher.objects.all().filter(status=True).count(),
    'total_course':models.Course.objects.all().count(),
    'total_question':models.Question.objects.all().count(),
    }
    return render(request,'exam/admin_dashboard.html',context=dict)

@login_required(login_url='adminlogin')
def admin_teacher_view(request):
    dict={
    'total_teacher':TMODEL.Teacher.objects.all().filter(status=True).count(),
    'pending_teacher':TMODEL.Teacher.objects.all().filter(status=False).count(),
    'salary':TMODEL.Teacher.objects.all().filter(status=True).aggregate(Sum('salary'))['salary__sum'],
    }
    return render(request,'exam/admin_teacher.html',context=dict)

@login_required(login_url='adminlogin')
def admin_view_teacher_view(request):
    teachers= TMODEL.Teacher.objects.all().filter(status=True)
    return render(request,'exam/admin_view_teacher.html',{'teachers':teachers})


@login_required(login_url='adminlogin')
def update_teacher_view(request,pk):
    teacher=TMODEL.Teacher.objects.get(id=pk)
    user=TMODEL.User.objects.get(id=teacher.user_id)
    userForm=TFORM.TeacherUserForm(instance=user)
    teacherForm=TFORM.TeacherForm(request.FILES,instance=teacher)
    mydict={'userForm':userForm,'teacherForm':teacherForm}
    if request.method=='POST':
        userForm=TFORM.TeacherUserForm(request.POST,instance=user)
        teacherForm=TFORM.TeacherForm(request.POST,request.FILES,instance=teacher)
        if userForm.is_valid() and teacherForm.is_valid():
            user=userForm.save()
            user.set_password(user.password)
            user.save()
            teacherForm.save()
            return redirect('admin-view-teacher')
    return render(request,'exam/update_teacher.html',context=mydict)



@login_required(login_url='adminlogin')
def delete_teacher_view(request,pk):
    teacher=TMODEL.Teacher.objects.get(id=pk)
    user=User.objects.get(id=teacher.user_id)
    user.delete()
    teacher.delete()
    return HttpResponseRedirect('/admin-view-teacher')




@login_required(login_url='adminlogin')
def admin_view_pending_teacher_view(request):
    teachers= TMODEL.Teacher.objects.all().filter(status=False)
    return render(request,'exam/admin_view_pending_teacher.html',{'teachers':teachers})


@login_required(login_url='adminlogin')
def approve_teacher_view(request,pk):
    teacherSalary=forms.TeacherSalaryForm()
    if request.method=='POST':
        teacherSalary=forms.TeacherSalaryForm(request.POST)
        if teacherSalary.is_valid():
            teacher=TMODEL.Teacher.objects.get(id=pk)
            teacher.salary=teacherSalary.cleaned_data['salary']
            teacher.status=True
            teacher.save()
        else:
            print("form is invalid")
        return HttpResponseRedirect('/admin-view-pending-teacher')
    return render(request,'exam/salary_form.html',{'teacherSalary':teacherSalary})

@login_required(login_url='adminlogin')
def reject_teacher_view(request,pk):
    teacher=TMODEL.Teacher.objects.get(id=pk)
    user=User.objects.get(id=teacher.user_id)
    user.delete()
    teacher.delete()
    return HttpResponseRedirect('/admin-view-pending-teacher')

@login_required(login_url='adminlogin')
def admin_view_teacher_salary_view(request):
    teachers= TMODEL.Teacher.objects.all().filter(status=True)
    return render(request,'exam/admin_view_teacher_salary.html',{'teachers':teachers})




@login_required(login_url='adminlogin')
def admin_student_view(request):
    dict={
    'total_student':SMODEL.Student.objects.all().count(),
    }
    return render(request,'exam/admin_student.html',context=dict)

@login_required(login_url='adminlogin')
def admin_view_student_view(request):
    students= SMODEL.Student.objects.all()
    return render(request,'exam/admin_view_student.html',{'students':students})



@login_required(login_url='adminlogin')
def update_student_view(request,pk):
    student=SMODEL.Student.objects.get(id=pk)
    user=SMODEL.User.objects.get(id=student.user_id)
    userForm=SFORM.StudentUserForm(instance=user)
    studentForm=SFORM.StudentForm(request.FILES,instance=student)
    mydict={'userForm':userForm,'studentForm':studentForm}
    if request.method=='POST':
        userForm=SFORM.StudentUserForm(request.POST,instance=user)
        studentForm=SFORM.StudentForm(request.POST,request.FILES,instance=student)
        if userForm.is_valid() and studentForm.is_valid():
            user=userForm.save()
            user.set_password(user.password)
            user.save()
            studentForm.save()
            return redirect('admin-view-student')
    return render(request,'exam/update_student.html',context=mydict)



# @login_required(login_url='adminlogin')
def delete_student_view(request,pk):
    student=SMODEL.Student.objects.get(id=pk)
    user=User.objects.get(id=student.user_id)
    user.delete()
    student.delete()
    return HttpResponseRedirect('/admin-view-student')


@login_required(login_url='adminlogin')
def admin_course_view(request):
    return render(request,'exam/admin_course.html')


@login_required(login_url='adminlogin')
def admin_add_course_view(request):
    courseForm=forms.CourseForm()
    if request.method=='POST':
        courseForm=forms.CourseForm(request.POST)
        if courseForm.is_valid():        
            courseForm.save()
        else:
            print("form is invalid")
        return HttpResponseRedirect('/admin-view-course')
    return render(request,'exam/admin_add_course.html',{'courseForm':courseForm})


@login_required(login_url='adminlogin')
def admin_view_course_view(request):
    courses = models.Course.objects.all()
    return render(request,'exam/admin_view_course.html',{'courses':courses})

@login_required(login_url='adminlogin')
def delete_course_view(request,pk):
    course=models.Course.objects.get(id=pk)
    course.delete()
    return HttpResponseRedirect('/admin-view-course')



@login_required(login_url='adminlogin')
def admin_question_view(request):
    return render(request,'exam/admin_question.html')


@login_required(login_url='adminlogin')
def admin_add_question_view(request):
    questionForm=forms.QuestionForm()
    if request.method=='POST':
        questionForm=forms.QuestionForm(request.POST)
        if questionForm.is_valid():
            question=questionForm.save(commit=False)
            course=models.Course.objects.get(id=request.POST.get('courseID'))
            question.course=course
            question.save()       
        else:
            print("form is invalid")
        return HttpResponseRedirect('/admin-view-question')
    return render(request,'exam/admin_add_question.html',{'questionForm':questionForm})


@login_required(login_url='adminlogin')
def admin_view_question_view(request):
    courses= models.Course.objects.all()
    return render(request,'exam/admin_view_question.html',{'courses':courses})

@login_required(login_url='adminlogin')
def view_question_view(request,pk):
    questions=models.Question.objects.all().filter(course_id=pk)
    return render(request,'exam/view_question.html',{'questions':questions})

@login_required(login_url='adminlogin')
def delete_question_view(request,pk):
    question=models.Question.objects.get(id=pk)
    question.delete()
    return HttpResponseRedirect('/admin-view-question')

@login_required(login_url='adminlogin')
def admin_view_student_marks_view(request):
    students= SMODEL.Student.objects.all()
    return render(request,'exam/admin_view_student_marks.html',{'students':students})

@login_required(login_url='adminlogin')
def admin_view_marks_view(request,pk):
    courses = models.Course.objects.all()
    response =  render(request,'exam/admin_view_marks.html',{'courses':courses})
    response.set_cookie('student_id',str(pk))
    return response

@login_required(login_url='adminlogin')
def admin_check_marks_view(request,pk):
    course = models.Course.objects.get(id=pk)
    student_id = request.COOKIES.get('student_id')
    student= SMODEL.Student.objects.get(id=student_id)

    results= models.Result.objects.all().filter(exam=course).filter(student=student)
    return render(request,'exam/admin_check_marks.html',{'results':results})
    




def aboutus_view(request):
    return render(request,'exam/aboutus.html')

def contactus_view(request):
    sub = forms.ContactusForm()
    if request.method == 'POST':
        sub = forms.ContactusForm(request.POST)
        if sub.is_valid():
            email = sub.cleaned_data['Email']
            name=sub.cleaned_data['Name']
            message = sub.cleaned_data['Message']
            send_mail(str(name)+' || '+str(email),message,settings.EMAIL_HOST_USER, settings.EMAIL_RECEIVING_USER, fail_silently = False)
            return render(request, 'exam/contactussuccess.html')
    return render(request, 'exam/contactus.html', {'form':sub})


@login_required(login_url='adminlogin')
def admin_shedule_course_view(request):
    if request.method == "POST":
        course_id = request.POST.get("course_id")
        start_time = request.POST.get("start_time")
        end_time = request.POST.get("end_time")

        try:
            course = models.Course.objects.get(id=course_id)
            course.start_time = start_time
            course.end_time = end_time
            course.save()
            # Optional: you can add a message or flash here
        except models.Course.DoesNotExist:
            pass  # Handle the error appropriately if needed

        send_simple_email(course_id)

        return redirect('admin-shedule-course')  # or render the same page with updated data

    courses = models.Course.objects.all()
    return render(request, 'exam/admin_shedule_course.html', {'courses': courses})


@login_required(login_url='adminlogin')
def admin_report_view(request):
    students = Student.objects.all()
    teachers = TMODEL.Teacher.objects.all()

    return render(request, 'exam/admin_report_main.html', {'students': students, 'teachers': teachers})

from django.http import HttpResponse
from reportlab.lib.pagesizes import A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet
from datetime import datetime
from .models import Student

def generate_monthly_student_report(request):
    now = datetime.now()
    current_month = now.month-1
    current_year = now.year

    current_month = int(request.GET.get('month'))
    current_year = int(request.GET.get('year'))

    # Filter students joined in current month
    students = Student.objects.filter(
        user__date_joined__year=current_year,
        user__date_joined__month=current_month
    )

    # Set up the PDF
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'inline; filename="monthly_student_report.pdf"'

    doc = SimpleDocTemplate(
        response,
        pagesize=A4,
        topMargin=30,  # Default is ~72; reduce to 30 or less
        bottomMargin=30,
        leftMargin=30,
        rightMargin=30
    )
    elements = []
    styles = getSampleStyleSheet()

    month_str = month_name[int(current_month)]  # Convert numeric month to name

    # Title
    title = Paragraph(f"<b>Student Report - {month_str} {current_year}</b>", styles['Title'])



    elements.append(title)
    elements.append(Spacer(1, 12))

    if students.exists():
        # Table Data (Headers + Rows)
        data = [['#', 'Full Name', 'Mobile', 'Qualification', 'Join Date']]
        for idx, student in enumerate(students, start=1):
            data.append([
                str(idx),
                student.user.get_full_name(),
                student.mobile,
                student.get_qualification_display(),
                student.user.date_joined.strftime('%d-%b-%Y')
            ])

        # Create the Table

        table = Table(data, repeatRows=1, colWidths=[30, 150, 100, 120, 90])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#4F81BD")),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
            ('BACKGROUND', (0, 1), (-1, -1), colors.whitesmoke),
            ('GRID', (0, 0), (-1, -1), 0.25, colors.grey),
        ]))
        elements.append(table)
    else:
        elements.append(Paragraph("No students joined this month.", styles['Normal']))

    doc.build(elements)
    return response

def generate_all_student_report(request):
    # Fetch all students
    students = Student.objects.select_related('user').all()

    # Set up the PDF response
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'inline; filename="all_students_report.pdf"'

    doc = SimpleDocTemplate(
        response,
        pagesize=A4,
        topMargin=30,
        bottomMargin=30,
        leftMargin=30,
        rightMargin=30
    )

    elements = []
    styles = getSampleStyleSheet()

    # Add title
    title = Paragraph("<b>All Students Report</b>", styles['Title'])
    elements.append(title)
    elements.append(Spacer(1, 12))

    if students.exists():
        # Table headers
        data = [['#', 'Full Name', 'Mobile', 'Qualification', 'Join Date']]

        # Add each student
        for idx, student in enumerate(students, start=1):
            data.append([
                str(idx),
                student.user.get_full_name(),
                student.mobile,
                student.get_qualification_display(),
                student.user.date_joined.strftime('%d-%b-%Y')
            ])

        # Define table with custom column widths
        table = Table(data, repeatRows=1, colWidths=[30, 150, 100, 120, 90])

        # Add styling
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#4F81BD")),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
            ('BACKGROUND', (0, 1), (-1, -1), colors.whitesmoke),
            ('GRID', (0, 0), (-1, -1), 0.25, colors.grey),
        ]))

        elements.append(table)
    else:
        elements.append(Paragraph("No students found.", styles['Normal']))

    doc.build(elements)
    return response




def monthly_teacher_report(request):
    now = datetime.now()
    current_month = now.month
    current_year = now.year

    current_month = int(request.GET.get('month'))
    current_year = int(request.GET.get('year'))



    teachers = TMODEL.Teacher.objects.filter(
        user__date_joined__year=current_year,
        user__date_joined__month=current_month
    )

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'inline; filename="monthly_teacher_report.pdf"'

    doc = SimpleDocTemplate(response, pagesize=A4, topMargin=30, bottomMargin=30, leftMargin=30, rightMargin=30)
    styles = getSampleStyleSheet()
    elements = []

    # elements.append(Paragraph(f"<b>Monthly Teacher Report - {current_month} {current_year}</b>", styles['Title']))

    month_str = month_name[int(current_month)]  # Convert numeric month to name
    elements.append(Paragraph(f"<b>Monthly Teacher Report - {month_str} {current_year}</b>", styles['Title']))

    elements.append(Spacer(1, 12))

    if teachers.exists():
        data = [['#', 'Full Name', 'Mobile', 'Salary', 'Join Date']]
        for idx, teacher in enumerate(teachers, start=1):
            data.append([
                str(idx),
                teacher.get_name,
                teacher.mobile,
                f"{teacher.salary} Rs" if teacher.salary else 'N/A',
                teacher.user.date_joined.strftime('%d-%b-%Y')
            ])

        table = Table(data, repeatRows=1, colWidths=[30, 150, 100, 80, 90])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#4BACC6")),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
            ('BACKGROUND', (0, 1), (-1, -1), colors.whitesmoke),
            ('GRID', (0, 0), (-1, -1), 0.25, colors.grey),

            # Alignments
            ('ALIGN', (0, 1), (1, -1), 'LEFT'),  # # and Full Name - left
            ('ALIGN', (2, 1), (2, -1), 'RIGHT'),  # Mobile - right
            ('ALIGN', (3, 1), (3, -1), 'RIGHT'),  # Salary - right
        ]))
        elements.append(table)
    else:
        elements.append(Paragraph("No teachers joined this month.", styles['Normal']))

    doc.build(elements)
    return response

def all_teachers_report(request):
    teachers = TMODEL.Teacher.objects.select_related('user').all()

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'inline; filename="all_teachers_report.pdf"'

    doc = SimpleDocTemplate(response, pagesize=A4, topMargin=30, bottomMargin=30, leftMargin=30, rightMargin=30)
    styles = getSampleStyleSheet()
    elements = []

    elements.append(Paragraph("<b>All Teachers Report</b>", styles['Title']))
    elements.append(Spacer(1, 12))

    if teachers.exists():
        data = [['#', 'Full Name', 'Mobile', 'Salary', 'Join Date']]
        for idx, teacher in enumerate(teachers, start=1):
            data.append([
                str(idx),
                teacher.get_name,
                teacher.mobile,
                f"{teacher.salary} Rs" if teacher.salary else 'N/A',
                teacher.user.date_joined.strftime('%d-%b-%Y')
            ])

        table = Table(data, repeatRows=1, colWidths=[30, 150, 100, 80, 90])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#4BACC6")),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
            ('BACKGROUND', (0, 1), (-1, -1), colors.whitesmoke),
            ('GRID', (0, 0), (-1, -1), 0.25, colors.grey),

            # Alignments
            ('ALIGN', (0, 1), (1, -1), 'LEFT'),  # # and Full Name - left
            ('ALIGN', (2, 1), (2, -1), 'RIGHT'),  # Mobile - right
            ('ALIGN', (3, 1), (3, -1), 'RIGHT'),  # Salary - right


        ]))
        elements.append(table)
    else:
        elements.append(Paragraph("No teachers found.", styles['Normal']))

    doc.build(elements)
    return response


from .models import Result

def generate_student_report(request):
    try:
        # Fetch student details
        student_id = request.GET.get('student_id')
        student = Student.objects.get(id=student_id)

        results = Result.objects.filter(student=student)

        # Set up PDF response
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = 'inline; filename="student_report.pdf"'

        doc = SimpleDocTemplate(response, pagesize=A4, topMargin=30, bottomMargin=30, leftMargin=30, rightMargin=30)
        styles = getSampleStyleSheet()
        elements = []

        # Title
        elements.append(Paragraph(f"<b>Student Report: {student.get_name}</b>", styles['Title']))
        elements.append(Spacer(1, 12))

        # Student Info Section
        student_info = [
            ['Name:', student.get_name],
            ['Qualification:', student.get_qualification_display()],
            ['Mobile:', student.mobile],
        ]
        student_info_table = Table(student_info, colWidths=[150, 320])
        student_info_table.setStyle(TableStyle([
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('GRID', (0, 0), (-1, -1), 0.25, colors.grey),
        ]))
        elements.append(student_info_table)
        elements.append(Spacer(1, 12))

        # Course Results Section
        if results.exists():
            data = [['#', 'Course Name', 'Marks Obtained', 'Date']]

            for idx, result in enumerate(results, start=1):
                data.append([
                    str(idx),
                    result.exam.course_name,
                    str(result.marks),
                    result.date.strftime('%d-%b-%Y')
                ])

            # Create Table for results
            result_table = Table(data, repeatRows=1, colWidths=[30, 250, 100, 100])
            result_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#4F81BD")),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 12),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
                ('BACKGROUND', (0, 1), (-1, -1), colors.whitesmoke),
                ('GRID', (0, 0), (-1, -1), 0.25, colors.grey),
            ]))
            elements.append(result_table)
        else:
            elements.append(Paragraph("No results found for this student.", styles['Normal']))

        # Build PDF
        doc.build(elements)
        return response

    except Student.DoesNotExist:
        return HttpResponse("Student not found.", status=404)




from .models import Course

def generate_all_exam_report(request):
    # Fetch all students and their results
    students = Student.objects.all()
    results = Result.objects.select_related('student', 'exam')

    # Set up PDF response
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'inline; filename="all_students_report_by_course.pdf"'

    doc = SimpleDocTemplate(response, pagesize=A4, topMargin=30, bottomMargin=30, leftMargin=30, rightMargin=30)
    styles = getSampleStyleSheet()
    elements = []

    # Title
    elements.append(Paragraph("<b>All Students Results Report </b>", styles['Title']))
    elements.append(Spacer(1, 12))

    # Get all unique courses
    courses = Course.objects.all()

    for course in courses:
        # Filter results for the current course
        course_results = results.filter(exam=course)

        if course_results.exists():
            # Title for the course
            elements.append(Paragraph(f"<b>Exam: {course.course_name}</b>", styles['Heading2']))
            elements.append(Spacer(1, 12))

            # Create table data for the current course
            data = [['#', 'Student Name', 'Marks Obtained', 'Date']]

            for idx, result in enumerate(course_results, start=1):
                data.append([
                    str(idx),
                    result.student.get_name,
                    str(result.marks),
                    result.date.strftime('%d-%b-%Y')
                ])

            # Create Table for the course
            result_table = Table(data, repeatRows=1, colWidths=[30, 250, 100, 100])
            result_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#4F81BD")),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 12),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
                ('BACKGROUND', (0, 1), (-1, -1), colors.whitesmoke),
                ('GRID', (0, 0), (-1, -1), 0.25, colors.grey),
            ]))

            elements.append(result_table)
            elements.append(Spacer(1, 12))  # Add space between courses

    # Build the PDF
    doc.build(elements)
    return response


from django.shortcuts import redirect

def generate_report_dispatcher(request):
    report_type = request.GET.get('type')

    if report_type == 'monthly_student':
        month = request.GET.get('month')
        year = request.GET.get('year')
        return redirect(f"{reverse('admin-report-month-student')}?month={month}&year={year}")

    elif report_type == 'monthly_teacher':
        month = request.GET.get('month')
        year = request.GET.get('year')
        return redirect(f"{reverse('admin-report-month-teacher')}?month={month}&year={year}")

    elif report_type == 'all_student':
        return redirect('admin-report-all-student')
    elif report_type == 'all_teacher':
        return redirect('admin-report-all-teacher')
    elif report_type == 'particular_student':
        student_id = request.GET.get('student_id')
        return redirect(f"{reverse('admin-report-student-info')}?student_id={student_id}")
    elif report_type == 'all_exam':
        return redirect('admin-report-exam-info')
    else:
        return redirect('admin-dashboard')  # Or show an error page




import openpyxl
import os
def send_simple_email(course_id):

    workbook = openpyxl.load_workbook("test_stuff2.xlsx")
    sheet_obj = workbook.active

    max_rows = sheet_obj.max_row
    print(max_rows)
    row = 0
    # Loop will print all columns name

    for i in range(2, max_rows + 1):
        cell_obj = sheet_obj.cell(row=i, column=1)
        if int(course_id) == int(cell_obj.value):
            row = i
            break
    print(row, "dsda")

    password=0

    if row != 0:
        cell_obj = sheet_obj.cell(row=row, column=2)
        print("sdgahd", cell_obj)
        print( str(cell_obj.value))
        password=str(cell_obj.value)

    students = SMODEL.Student.objects.all()
    usernames = [s.user.username for s in students]
    print(usernames)

    course = models.Course.objects.get(id=course_id)

    subject = f"New Test Added: {course.course_name} ({course.start_time.strftime('%d %b %Y %I:%M %p')} - {course.end_time.strftime('%d %b %Y %I:%M %p')})"


    message = (
        f"Dear Student,\n\n"
        f"A new test has been added and Exam name is: {course.course_name}.\n\n"
        f"🕒 Test Timing:\n"
        f"Start: {course.start_time.strftime('%d %b %Y %I:%M %p')}\n"
        f"End:   {course.end_time.strftime('%d %b %Y %I:%M %p')}\n\n"
        f"📋 Test Details:\n"
        f"Total Questions: {course.question_number}\n"
        f"Total Marks:     {course.total_marks}\n"
        f"Negative Marking: {'Yes' if course.negative_marking else 'No'}\n\n"
        f"🔐 Password to access the test: {password}\n\n"
        f"Please be on time and ensure a stable internet connection.\n\n"
        f"Best Regards,\n"
        f"Examination Department"
    )

    from_email = settings.DEFAULT_FROM_EMAIL
    # recipient_list = ['patilap1854@gmail.com','as31563156@gmail.com']

    recipient_list=usernames
    # test= SMODEL.Student.objects.all()

    send_mail(subject, message, from_email, recipient_list)




